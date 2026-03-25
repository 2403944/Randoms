#!/usr/bin/env python3
"""
Fix: Create a data/ folder for all Excel files and update all path references.
Run: python fix_data_folder.py

After running this:
  1. Put all your Excel files in genai_monitor/data/
  2. Put .env in genai_monitor/
  3. cd genai_monitor && python run_all.py

Folder structure after fix:
  genai_monitor/
  ├── .env
  ├── app.py
  ├── run_all.py
  ├── data/
  │   ├── Metrics_template02.xlsx
  │   ├── final_eval_results01.xlsx
  │   ├── agent_query_augmentations.xlsx
  │   └── Augmentations 2.xlsx
  ├── routes/
  ├── tabs/
  └── templates/
"""
import os

BASE = "genai_monitor"

def w(path, content):
    full = os.path.join(BASE, path)
    os.makedirs(os.path.dirname(full), exist_ok=True)
    with open(full, "w", encoding="utf-8") as f:
        f.write(content)
    print(f"  Updated: {path}")

def build():
    print(f"Creating data/ folder and updating all file paths...\n")

    # Create data folder
    os.makedirs(os.path.join(BASE, "data"), exist_ok=True)
    print(f"  Created: data/")

    # ==================== data_paths.py ====================
    # Single source of truth for ALL file paths
    w("data_paths.py", r'''"""
Single source of truth for all data file paths.
All Excel/data files live in the data/ subfolder.
.env stays in the project root (genai_monitor/).
"""
from pathlib import Path

# Project root = genai_monitor/
PROJECT_ROOT = Path(__file__).resolve().parent

# Data folder = genai_monitor/data/
DATA_DIR = PROJECT_ROOT / "data"

# Excel file paths
METRICS_TEMPLATE_PATH = DATA_DIR / "Metrics_template02.xlsx"
EVAL_RESULTS_PATH = DATA_DIR / "final_eval_results01.xlsx"
AGENT_AUGMENTATIONS_PATH = DATA_DIR / "agent_query_augmentations.xlsx"
AUGMENTATIONS_EXCEL_PATH = DATA_DIR / "Augmentations 2.xlsx"

# .env stays in project root
ENV_PATH = PROJECT_ROOT / ".env"
''')

    # ==================== Update techniques_loader.py ====================
    w("techniques_loader.py", r'''"""Loads augmentation techniques from Excel file or falls back to defaults."""
import pandas as pd
from data_paths import AUGMENTATIONS_EXCEL_PATH
from prompts import AUGMENTATION_PROMPTS

try:
    from loguru import logger
except ImportError:
    import logging
    logger = logging.getLogger(__name__)


def load_augmentation_techniques_from_excel():
    techniques_path = AUGMENTATIONS_EXCEL_PATH
    result = []
    try:
        if not techniques_path.exists():
            for name, prompt in AUGMENTATION_PROMPTS.items():
                result.append({"category": "General", "sub_category": "Text Transformation", "name": name, "description": prompt.split("\n\nText:")[0]})
            return result
        xls = pd.ExcelFile(techniques_path)
        for sheet_name in xls.sheet_names:
            category = str(sheet_name).strip()
            try:
                df = pd.read_excel(xls, sheet_name=sheet_name)
                df.columns = [str(c).strip() for c in df.columns]
                sub_cat_col = name_col = desc_col = None
                for c in df.columns:
                    cl = c.lower().replace("_", " ").strip()
                    if "sub" in cl and "cat" in cl:
                        sub_cat_col = c
                    elif cl in ("name", "technique", "technique name"):
                        name_col = c
                    elif "desc" in cl:
                        desc_col = c
                if not name_col:
                    remaining = [c for c in df.columns if c != sub_cat_col and c != desc_col]
                    name_col = remaining[0] if remaining else None
                for _, row in df.iterrows():
                    name = str(row.get(name_col, "")).strip() if name_col else ""
                    if not name or name.lower() == "nan":
                        continue
                    sub_cat = str(row.get(sub_cat_col, "")).strip() if sub_cat_col else ""
                    if not sub_cat or sub_cat.lower() == "nan":
                        sub_cat = "General"
                    desc = str(row.get(desc_col, "")).strip() if desc_col else ""
                    if desc.lower() == "nan":
                        desc = ""
                    result.append({"category": category, "sub_category": sub_cat, "name": name, "description": desc})
            except Exception as e:
                logger.error(f"Error reading sheet '{sheet_name}': {e}")
        if not result:
            for name, prompt in AUGMENTATION_PROMPTS.items():
                result.append({"category": "General", "sub_category": "Text Transformation", "name": name, "description": prompt.split("\n\nText:")[0]})
    except Exception as e:
        logger.error(f"Error loading augmentation techniques: {e}")
        for name, prompt in AUGMENTATION_PROMPTS.items():
            result.append({"category": "General", "sub_category": "Text Transformation", "name": name, "description": prompt.split("\n\nText:")[0]})
    return result


ALL_AUGMENTATION_TECHNIQUES = load_augmentation_techniques_from_excel()
''')

    # ==================== Update report_generator.py ====================
    w("report_generator.py", r'''"""
Main ReportGenerator class - FULLY SELF-CONTAINED.
Uses data_paths.py for all file locations.
"""
import json
import pandas as pd
from datetime import datetime

try:
    from loguru import logger
except ImportError:
    import logging
    logger = logging.getLogger(__name__)

from config import get_default_config
from data_paths import METRICS_TEMPLATE_PATH, EVAL_RESULTS_PATH, AGENT_AUGMENTATIONS_PATH
from utils import (find_column, safe_numeric_conversion, normalize_text, text_equal, clean_text, normalize_query_match)
from tabs.metrics_summary import generate_metrics_summary_table
from tabs.data_assurance import create_test_coverage_sunburst, generate_augmented_data_table, get_coverage_edit_data
from tabs.model_quality_assurance import (generate_interactive_details_table, create_metrics_bar_chart, create_overall_pie_chart, create_score_comparison_chart, create_disaggregated_table, load_recovery_loop_global)
from tabs.secondary_llm import generate_secondary_llm_table


class ReportGenerator:
    def __init__(self, config=None):
        logger.info('Initializing ReportGenerator')
        self.config = config or get_default_config()
        self.report_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.metric_details_excel_path = EVAL_RESULTS_PATH

    def _get_chart_layout(self, theme):
        return {'font': {'family': theme['font_family'], 'size': 12}, 'plot_bgcolor': 'rgba(0,0,0,0)', 'paper_bgcolor': 'rgba(0,0,0,0)', 'height': self.config['chart_height'], 'margin': dict(l=60, r=60, t=80, b=100), 'xaxis': {'showgrid': True, 'gridcolor': theme['border_color'], 'tickangle': -45, 'tickfont': {'size': 10}}, 'yaxis': {'showgrid': True, 'gridcolor': theme['border_color'], 'tickfont': {'size': 10}}}

    def _calculate_status(self, df, metrics, thresholds):
        try:
            df_copy = df.copy()
            for metric in metrics:
                if metric not in df_copy.columns: continue
                status_col = f"{metric}_status"; reason_col = f"{metric}-reason"; threshold = thresholds.get(metric)
                if threshold is None: df_copy[status_col] = "Unknown"; continue
                if reason_col in df_copy.columns:
                    df_copy[status_col] = df_copy.apply(lambda row, m=metric, rc=reason_col, th=threshold: self._determine_status_with_reason(row, m, rc, th), axis=1)
                else:
                    df_copy[status_col] = df_copy[metric].apply(lambda x, m=metric, th=threshold: self._determine_status_simple(x, m, th))
            return df_copy
        except Exception as e: logger.error(f"Error calculating status: {e}"); return df

    def _determine_status_with_reason(self, row, metric, reason_col, threshold):
        reason = str(row.get(reason_col, "")).strip().upper()
        if reason == "NA": return "Skipped"
        score = safe_numeric_conversion(row[metric])
        if pd.isna(score): return "Failed"
        if metric in self.config['reverse_metrics']: return "Passed" if score <= threshold else "Failed"
        return "Passed" if score >= threshold else "Failed"

    def _determine_status_simple(self, value, metric, threshold):
        score = safe_numeric_conversion(value)
        if pd.isna(score): return "Failed"
        if metric in self.config['reverse_metrics']: return "Passed" if score <= threshold else "Failed"
        return "Passed" if score >= threshold else "Failed"

    def _prepare_modal_data(self, df, metrics, thresholds):
        try:
            modal_data = {}; metric_sheets = {}
            try:
                if self.metric_details_excel_path.exists():
                    xls = pd.ExcelFile(self.metric_details_excel_path)
                    for sheet in xls.sheet_names:
                        try: metric_sheets[sheet] = pd.read_excel(xls, sheet_name=sheet)
                        except: continue
            except Exception as e: logger.warning(f"Unable to read metric details excel: {e}")
            normalized_sheet_map = {str(s).strip().lower(): s for s in metric_sheets.keys()}
            for idx, row in df.iterrows():
                modal_data[idx] = {'query': clean_text(row.get('Query', 'N/A')), 'response': clean_text(row.get('Response', 'N/A')), 'timestamp': self.report_timestamp, 'metrics': {}}
                modal_data[idx]["metric_fields"] = {}
                for metric in metrics:
                    norm_metric = str(metric).strip().lower()
                    if norm_metric in normalized_sheet_map:
                        sheet_name = normalized_sheet_map[norm_metric]; metric_df = metric_sheets.get(sheet_name)
                        if isinstance(metric_df, pd.DataFrame):
                            query_col = find_column(metric_df, ['Query','query','Question']); found_entry = {}
                            if query_col and query_col in metric_df.columns:
                                metric_df[query_col] = metric_df[query_col].ffill(); details_query = normalize_text(row.get('Query', ''))
                                mask = metric_df[query_col].apply(lambda x: normalize_text(x) == details_query); matched = metric_df[mask]
                                if not matched.empty:
                                    found_entry = matched.iloc[0].replace({pd.NA: None}).to_dict() if len(matched)==1 else [r.replace({pd.NA: None}).to_dict() for _, r in matched.iterrows()]
                                else:
                                    mask2 = metric_df[query_col].apply(lambda x: text_equal(x, details_query)); matched2 = metric_df[mask2]
                                    if not matched2.empty: found_entry = matched2.iloc[0].replace({pd.NA: None}).to_dict() if len(matched2)==1 else [r.replace({pd.NA: None}).to_dict() for _, r in matched2.iterrows()]
                            modal_data[idx]["metric_fields"][str(metric).strip()] = found_entry
                        else: modal_data[idx]["metric_fields"][metric] = {}
                    else: modal_data[idx]["metric_fields"][metric] = {}
                for metric in metrics:
                    ms = str(metric).strip(); mn = normalize_text(ms)
                    md = {'score': clean_text(row.get(metric, 'N/A')), 'threshold': str(thresholds.get(metric, 'N/A')), 'status': clean_text(row.get(f"{metric}_status", 'Unknown')), 'additional_fields': {}}
                    for col in row.index:
                        cn = normalize_text(str(col).strip())
                        if cn == mn or cn.endswith("status"): continue
                        if cn.startswith(mn):
                            fn = cn.replace(mn,"").strip(" _-"); fv = clean_text(row.get(col, ''))
                            if fv != 'N/A': md['additional_fields'][fn] = fv
                    modal_data[idx]['metrics'][ms] = md
            return modal_data
        except Exception as e: logger.error(f"Error preparing modal data: {e}"); return {}

    def _create_modern_charts(self, summary_df, overall_df, metrics_df, metrics_col, score_col, threshold_col):
        try:
            charts = {}; layout_fn = self._get_chart_layout
            charts['metrics'] = create_metrics_bar_chart(summary_df, self.config, layout_fn)
            charts['overall'] = create_overall_pie_chart(overall_df, self.config, layout_fn)
            charts['comparison'] = create_score_comparison_chart(metrics_df, metrics_col, score_col, threshold_col, self.config, layout_fn)
            charts['coverage'] = create_test_coverage_sunburst(excel_path=METRICS_TEMPLATE_PATH, sheet_name='Test data coverage')
            charts['augmentation'] = generate_augmented_data_table(excel_path=AGENT_AUGMENTATIONS_PATH, sheet_name='Sheet1')
            charts['disaggregated'] = create_disaggregated_table(METRICS_TEMPLATE_PATH)
            return charts
        except Exception as e: logger.error(f"Error creating charts: {e}"); return {}

    def generate_report(self, metrics_df, details_df, summary_df, overall_df):
        try:
            logger.info("Starting report generation")
            if any(df.empty for df in [metrics_df, details_df, summary_df, overall_df]):
                return self._generate_error_report("One or more input DataFrames are empty")
            mc = find_column(metrics_df, ['Metrics','Metric','metric']); tc = find_column(metrics_df, ['Threshold_Value','Threshold','threshold']); sc = find_column(metrics_df, ['aggregate_score','score','Score'])
            if not all([mc, tc, sc]): return self._generate_error_report("Required columns not found")
            all_metrics = [m for m in metrics_df[mc].dropna().astype(str).tolist() if m.strip().lower() != 'knowledge retention']
            thresholds = {k: v for k, v in zip(metrics_df[mc], metrics_df[tc]) if str(k).strip().lower() != 'knowledge retention'}
            details_df = self._calculate_status(details_df, all_metrics, thresholds)
            charts = self._create_modern_charts(summary_df, overall_df, metrics_df, mc, sc, tc)
            modal_data = self._prepare_modal_data(details_df, all_metrics, thresholds)
            from templates.html_report import generate_modern_html
            return generate_modern_html(config=self.config, report_timestamp=self.report_timestamp, metrics_df=metrics_df, details_df=details_df, all_metrics=all_metrics, charts=charts, modal_data=modal_data, metric_details_excel_path=self.metric_details_excel_path)
        except Exception as e: return self._generate_error_report(f"Error: {e}")

    def _generate_error_report(self, msg):
        return f"<!DOCTYPE html><html><body style='font-family:Arial;padding:40px;text-align:center;'><h1>Report Generation Error</h1><div style='color:#dc3545;margin:20px;'>{msg}</div></body></html>"
''')

    # ==================== Update routes/main_routes.py ====================
    w("routes/main_routes.py", r'''"""Main routes - demo report page and run comparison."""
import pandas as pd
from flask import Blueprint, request, jsonify
from data_paths import METRICS_TEMPLATE_PATH
from config import get_default_config

main_bp = Blueprint('main', __name__)


@main_bp.route("/")
def demo_report():
    try:
        from report_generator import ReportGenerator
        fp = str(METRICS_TEMPLATE_PATH)
        metrics_df = pd.read_excel(fp, sheet_name='Metrics Interpretability')
        details_df = pd.read_excel(fp, sheet_name='Test Data')
        summary_df = pd.read_excel(fp, sheet_name='Metrics_wise Pass-Fail')
        overall_df = pd.read_excel(fp, sheet_name='Overall Summary')
        rg = ReportGenerator()
        html = rg.generate_report(metrics_df, details_df, summary_df, overall_df)
        return html
    except Exception as e:
        return f"<h1>Error</h1><pre>{e}</pre>", 500


@main_bp.route("/compare-runs", methods=["POST"])
def compare_runs():
    try:
        from tabs.model_quality_assurance import create_run_comparison_chart
        pf = request.files.get("previous_run")
        cf = request.files.get("current_run")
        if not pf or not cf:
            return "<div><b>Upload both files</b></div>"
        config = get_default_config()
        return create_run_comparison_chart(pd.read_excel(pf), pd.read_excel(cf), config["theme"])
    except Exception as e:
        return f"<div><b>Error: {e}</b></div>"


@main_bp.route("/update-coverage", methods=["POST"])
def update_coverage():
    try:
        payload = request.get_json()
        updates = payload.get("updates", {})
        excel_path = METRICS_TEMPLATE_PATH
        df = pd.read_excel(excel_path, sheet_name='Test data coverage')
        if "expected_record_count" not in df.columns:
            df["expected_record_count"] = None
        df["topic"] = df["topic"].astype(str).str.strip().str.upper()
        df["sub_topic"] = df["sub_topic"].astype(str).str.strip().str.title()
        df["sub_subtopic"] = df["sub_subtopic"].astype(str).str.strip().replace({"nan": None, "None": None, "": None})
        agg = df.groupby(["topic", "sub_topic", "sub_subtopic"], dropna=False).agg(count=("query", lambda x: x.notna().sum())).reset_index()
        for ri_str, cols in updates.items():
            ri = int(ri_str)
            if ri >= len(agg): continue
            ar = agg.iloc[ri]
            nv = cols.get("expected_record_count", "")
            nv = int(nv) if str(nv).strip().isdigit() else None
            mask = ((df["topic"] == ar["topic"]) & (df["sub_topic"] == ar["sub_topic"]) & (df["sub_subtopic"].fillna("-") == (ar["sub_subtopic"] if pd.notna(ar["sub_subtopic"]) else "-")))
            df.loc[mask, "expected_record_count"] = nv
        from openpyxl import load_workbook
        wb = load_workbook(excel_path)
        ws_name = 'Test data coverage'
        if ws_name in wb.sheetnames: del wb[ws_name]
        ws = wb.create_sheet(ws_name)
        for ci, col in enumerate(df.columns, 1): ws.cell(row=1, column=ci, value=col)
        for ri, row in enumerate(df.itertuples(index=False), 2):
            for ci, val in enumerate(row, 1): ws.cell(row=ri, column=ci, value=val)
        wb.save(excel_path)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})
''')

    # ==================== Update tabs/data_assurance.py - fix paths ====================
    w("tabs/data_assurance.py", r'''"""Data Assurance tab - coverage sunburst, augmentation pipeline, golden dataset."""
import json
import pandas as pd
import plotly.express as px
import plotly.offline as plot

try:
    from loguru import logger
except ImportError:
    import logging
    logger = logging.getLogger(__name__)

from utils import highlight_diff
from data_paths import METRICS_TEMPLATE_PATH


def create_test_coverage_sunburst(excel_path, sheet_name):
    try:
        df = pd.read_excel(excel_path, sheet_name=sheet_name)
        df["topic"] = df["topic"].astype(str).str.strip().str.upper()
        df["sub_topic"] = df["sub_topic"].astype(str).str.strip().str.title()
        df["sub_subtopic"] = df["sub_subtopic"].astype(str).str.strip().replace({"": None, "nan": None, "None": None})
        has_expected = "expected_record_count" in df.columns
        agg_cols = {"count": ("query", lambda x: x.notna().sum())}
        if has_expected:
            agg_cols["expected"] = ("expected_record_count", "sum")
        sunburst_df = df.groupby(["topic", "sub_topic", "sub_subtopic"], dropna=False).agg(**agg_cols).reset_index()
        table_rows = ""
        for _, row in sunburst_df.iterrows():
            actual = int(row["count"])
            sub_sub = row["sub_subtopic"] if pd.notna(row["sub_subtopic"]) else "-"
            expected_val = int(row["expected"]) if has_expected and pd.notna(row.get("expected")) else "-"
            table_rows += f"<tr><td>{row['topic']}</td><td>{row['sub_topic']}</td><td>{sub_sub}</td><td>{expected_val}</td><td>{actual}</td></tr>"
        table_html = f'<div style="width:100%;margin-bottom:32px;"><div class="table-container" style="width:100%;max-width:100%;overflow-x:auto;"><table class="modern-table" style="width:100%;"><thead><tr><th>Topic</th><th>Sub Topic</th><th>Sub Subtopic</th><th>Expected Record Count</th><th>Actual Record Count</th></tr></thead><tbody>{table_rows}</tbody></table></div></div>'
        fig = px.sunburst(sunburst_df, path=["topic", "sub_topic", "sub_subtopic"], values="count", color="topic", branchvalues="total", maxdepth=2)
        fig.update_traces(texttemplate="<b>%{label}</b><br>Count: %{value}<br>%{percentRoot:.0%}", hovertemplate="<b>%{label}</b><br>Count: %{value}<br>%{percentRoot:.0%}<extra></extra>", textfont=dict(size=14))
        fig.update_layout(sunburstcolorway=px.colors.qualitative.Pastel, margin=dict(t=60, l=20, r=20, b=20))
        chart_html = plot.plot(fig, output_type="div", include_plotlyjs=False)
        return table_html + f'<div style="display:flex;justify-content:center;width:100%;"><div style="width:fit-content;">{chart_html}</div></div>'
    except Exception as e:
        return "<div class='text-center p-3'><b>Error loading test coverage</b></div>"


def generate_augmented_data_table(excel_path, sheet_name):
    try:
        df = pd.read_excel(excel_path, sheet_name=sheet_name)
        if df.empty:
            return "<div class='text-center p-3'>No augmentation data available</div>"
        df.columns = [str(c).strip() for c in df.columns]
        base_col = aug_col = case_col = None
        for c in df.columns:
            nm = c.lower()
            if "base" in nm and "query" in nm: base_col = c
            if "augmented" in nm: aug_col = c
            if nm.strip() == "case": case_col = c
        display_cols = [c for c in df.columns if case_col is None or c != case_col]
        case_styles = {"P": ("P", "#1565C0", "#E3F2FD"), "N": ("N", "#B71C1C", "#FFEBEE"), "E": ("E", "#1B5E20", "#E8F5E9")}
        html = ['<div class="table-container"><table class="modern-table"><thead><tr>']
        for col in display_cols:
            html.append(f'<th>{col.replace("_", " ").title()}</th>')
        html.append('</tr></thead><tbody>')
        for _, row in df.iterrows():
            html.append('<tr>')
            bq = str(row[base_col]) if base_col and pd.notna(row.get(base_col)) else ""
            aq = str(row[aug_col]) if aug_col and pd.notna(row.get(aug_col)) else ""
            badge = ""
            if case_col and pd.notna(row.get(case_col)):
                raw = str(row[case_col]).strip().lower()
                key = "P" if raw.startswith("p") else ("N" if raw.startswith("n") else ("E" if raw.startswith("e") else ""))
                if key in case_styles:
                    lb, fg, bg = case_styles[key]
                    badge = f"<span style='display:inline-block;padding:2px 8px;border-radius:4px;font-size:0.75rem;font-weight:800;background:{bg};color:{fg};border:1px solid {fg};'>{lb}</span>"
            for col in display_cols:
                val = row.get(col, "")
                if col == aug_col:
                    dh = highlight_diff(bq, aq)
                    if badge:
                        cc = f"<div style='position:relative;min-height:24px;'><div style='position:absolute;top:0;right:0;'>{badge}</div><div style='padding-right:36px;'>{dh}</div></div>"
                    else:
                        cc = dh
                    html.append(f'<td>{cc}</td>')
                else:
                    html.append(f'<td>{"N/A" if pd.isna(val) else str(val)}</td>')
            html.append('</tr>')
        html.append('</tbody></table></div>')
        return '\n'.join(html)
    except Exception as e:
        return f"<div class='alert alert-error'>Error: {e}</div>"


def get_coverage_edit_data():
    try:
        excel_path = METRICS_TEMPLATE_PATH
        df = pd.read_excel(excel_path, sheet_name='Test data coverage')
        df["topic"] = df["topic"].astype(str).str.strip().str.upper()
        df["sub_topic"] = df["sub_topic"].astype(str).str.strip().str.title()
        df["sub_subtopic"] = df["sub_subtopic"].astype(str).str.strip().replace({"nan": None, "None": None, "": None})
        has_expected = "expected_record_count" in df.columns
        agg_cols = {"count": ("query", lambda x: x.notna().sum())}
        if has_expected:
            agg_cols["expected_record_count"] = ("expected_record_count", "first")
        agg = df.groupby(["topic", "sub_topic", "sub_subtopic"], dropna=False).agg(**agg_cols).reset_index()
        rows = []
        for _, row in agg.iterrows():
            rows.append({
                "topic": row["topic"], "sub_topic": row["sub_topic"],
                "sub_subtopic": row["sub_subtopic"] if pd.notna(row["sub_subtopic"]) else "-",
                "expected_record_count": int(row["expected_record_count"]) if has_expected and pd.notna(row.get("expected_record_count")) else "",
                "actual_record_count": int(row["count"]),
            })
        return rows
    except Exception:
        return []


def generate_augmentation_pipeline_html():
    return """<div>
<div class="wizard-steps" id="wizardSteps"><div class="wizard-step active" onclick="goToWizardStep(1)"><div class="wizard-step-num">1</div><div class="wizard-step-label">Dataset Upload</div></div><div class="wizard-connector" id="wconn1"></div><div class="wizard-step" onclick="goToWizardStep(2)"><div class="wizard-step-num">2</div><div class="wizard-step-label">Data Coverage</div></div><div class="wizard-connector" id="wconn2"></div><div class="wizard-step" onclick="goToWizardStep(3)"><div class="wizard-step-num">3</div><div class="wizard-step-label">Perform Augmentation</div></div><div class="wizard-connector" id="wconn3"></div><div class="wizard-step" onclick="goToWizardStep(4)"><div class="wizard-step-num">4</div><div class="wizard-step-label">Review & Download</div></div></div>
<div class="wizard-body" id="wizStep1" style="display:block;"><div class="card"><div class="card-header"><h3>Upload Your Dataset</h3></div><div class="card-content"><p style="color:#555;margin-bottom:16px;">Upload a CSV or Excel file containing two columns: <strong>input</strong> and <strong>Expected output</strong>.</p><div class="upload-zone" id="uploadZone" onclick="document.getElementById('datasetFile').click();" ondragover="event.preventDefault();this.classList.add('dragover');" ondragleave="this.classList.remove('dragover');" ondrop="event.preventDefault();this.classList.remove('dragover');handleFileDrop(event);"><div style="font-size:1.1rem;font-weight:600;color:#333;">Click or drag &amp; drop your file here</div><div style="font-size:0.85rem;color:#888;margin-top:6px;">Supports .csv, .xlsx, .xls</div></div><input type="file" id="datasetFile" accept=".csv,.xlsx,.xls" style="display:none;" onchange="handleFileUpload(this)"><div id="uploadStatus" style="margin-top:12px;"></div><div id="uploadPreview" class="preview-table" style="display:none;margin-top:16px;"></div></div></div><div class="wizard-actions"><div></div><button class="btn btn-primary" id="step1Next" disabled onclick="goToWizardStep(2)">Next &#8594;</button></div></div>
<div class="wizard-body" id="wizStep2" style="display:none;"><div class="card"><div class="card-header"><h3>Define Data Coverage Distribution</h3></div><div class="card-content"><p style="color:#555;margin-bottom:16px;">Specify the expected percentage for each topic / sub-topic / sub-sub-topic. Percentages must sum to <strong>100%</strong>.</p><div id="coverageRows"></div><button class="btn btn-secondary" onclick="addCoverageRow()" style="margin-top:10px;">+ Add Row</button><div class="coverage-total" id="coverageTotalDisplay">Total: 0%</div><div id="coverageSaveStatus" style="margin-top:10px;"></div></div></div><div class="wizard-actions"><button class="btn btn-secondary" onclick="goToWizardStep(1)">&#8592; Back</button><button class="btn btn-primary" id="step2Next" disabled onclick="saveCoveragePlan()">Save & Next &#8594;</button></div></div>
<div class="wizard-body" id="wizStep3" style="display:none;"><div class="card"><div class="card-header"><h3>Perform Augmentation</h3></div><div class="card-content">
<div style="margin-bottom:20px;"><div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px;"><label style="font-weight:600;">Select Input Rows</label><label style="display:flex;align-items:center;gap:6px;cursor:pointer;font-weight:600;font-size:0.85rem;color:#555;"><input type="checkbox" id="selectAllInputs" onchange="toggleAllInputs(this.checked)" style="accent-color:#2E86AB;width:16px;height:16px;cursor:pointer;">Select All</label></div><div class="aug-multi-select" id="inputSelectBox" onclick="toggleAugDropdown('inputDropdown')"><div class="chips" id="inputChips"></div><div style="font-size:0.85rem;color:#aaa;padding:4px;" id="inputPlaceholder">Click to select inputs...</div><div class="dropdown" id="inputDropdown"></div></div></div>
<div style="margin-bottom:20px;"><label style="font-weight:600;display:block;margin-bottom:12px;font-size:1.05rem;">Augmentation Techniques</label><div style="display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:12px;margin-bottom:16px;"><div style="display:flex;align-items:center;gap:8px;flex-wrap:wrap;"><div style="display:flex;gap:4px;border:2px solid #e2e8f0;border-radius:14px;overflow:hidden;padding:3px;" id="augTechTabs"><button class="aug-tech-tab active" onclick="filterAugTechs('all',this)">All</button></div><label style="display:flex;align-items:center;gap:6px;cursor:pointer;font-weight:600;font-size:0.85rem;color:#555;white-space:nowrap;margin-left:8px;"><input type="checkbox" id="augTechSelectAll" onchange="toggleAllAugTechs(this.checked)" style="accent-color:#2E86AB;width:16px;height:16px;cursor:pointer;">Select All</label></div><input type="text" id="augTechSearch" placeholder="Search techniques..." oninput="filterAugTechs(null,null)" style="padding:8px 14px;border:1px solid #ccc;border-radius:50px;width:220px;font-size:0.88rem;"></div><div id="augTechGrid" style="display:grid;grid-template-columns:repeat(auto-fill,minmax(190px,1fr));gap:8px;max-height:340px;overflow-y:auto;padding:4px;margin-bottom:12px;"></div><div id="augTechCount" style="font-size:0.85rem;font-weight:600;color:#555;margin-bottom:8px;"></div><div class="chips" id="techChips" style="margin-bottom:8px;"></div><div id="variationCountsArea" style="margin-top:14px;"></div></div>
<div style="display:flex;align-items:center;gap:16px;flex-wrap:wrap;">
<button class="btn btn-primary" id="augmentBtn" onclick="performAugmentation()" disabled>Run Augmentation</button>
<button class="btn btn-secondary" onclick="openCustomAugModal()" style="border:2px solid #A23B72;color:#A23B72;font-weight:700;">Custom Augmentation</button>
<div id="agenticConfigBadge" style="display:none;padding:6px 14px;background:#f8f0fc;border:1px solid #e2d0f0;border-radius:8px;font-size:0.85rem;color:#A23B72;font-weight:600;"><span id="agenticBadgeCount">agentic config(s) saved</span> <span onclick="agenticConfigs=[];this.parentElement.style.display='none';" style="cursor:pointer;margin-left:8px;color:#dc3545;">&#10005;</span></div>
</div>
<div id="augProgress" class="aug-progress" style="display:none;"><div class="aug-progress-bar"><div class="aug-progress-fill" id="augProgressFill" style="width:0%"></div></div><div class="aug-progress-text" id="augProgressText">Processing...</div></div>
<div id="augPreviewToggle" style="display:none;margin-top:16px;margin-bottom:10px;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:10px;"><div style="display:inline-flex;border:2px solid #e2e8f0;border-radius:8px;overflow:hidden;"><button id="toggleRaw" onclick="setPreviewMode('raw')" style="padding:8px 20px;border:none;cursor:pointer;font-weight:600;font-size:0.88rem;background:#2E86AB;color:white;">RAW</button><button id="toggleDiff" onclick="setPreviewMode('diff')" style="padding:8px 20px;border:none;cursor:pointer;font-weight:600;font-size:0.88rem;background:white;color:#555;">Diff Mode</button></div><select id="augResultTypeSelect" onchange="renderAugResultPreview()" style="padding:8px 14px;border:2px solid #A23B72;border-radius:8px;font-size:0.88rem;font-weight:600;color:#A23B72;cursor:pointer;background:white;"><option value="nonagentic">Non-Agentic Results</option><option value="agentic">Agentic Results</option></select></div>
<div id="augResultPreview"></div>
</div></div><div class="wizard-actions"><button class="btn btn-secondary" onclick="goToWizardStep(2)">&#8592; Back</button><button class="btn btn-primary" id="step3Next" disabled onclick="goToWizardStep(4)">Review Results &#8594;</button></div></div>
<div class="wizard-body" id="wizStep4" style="display:none;"><div class="card"><div class="card-header" style="display:flex;justify-content:space-between;align-items:center;"><h3>&#10004; Review Augmented Dataset</h3><div style="display:flex;gap:10px;align-items:center;"><div id="goldenSaveStatus" style="font-size:0.9rem;"></div><select id="reviewTypeSelect" onchange="renderReviewTable()" style="padding:6px 12px;border:2px solid #A23B72;border-radius:6px;font-size:0.82rem;font-weight:600;color:#A23B72;cursor:pointer;background:white;"><option value="nonagentic">Non-Agentic</option><option value="agentic">Agentic</option></select><button class="btn btn-success" onclick="downloadAugmented()">&#11015; Download CSV</button><button class="btn btn-primary" onclick="saveAsGolden()">&#128190; Save as Golden Dataset</button></div></div><div class="card-content"><div id="reviewTable" style="max-height:500px;overflow:auto;"></div></div></div><div class="wizard-actions"><button class="btn btn-secondary" onclick="goToWizardStep(3)">&#8592; Back</button><div></div></div></div>
</div>"""


def generate_golden_dataset_html():
    return """
        <div class="card">
            <div class="card-header" style="display:flex;justify-content:space-between;align-items:center;">
                <h3>Golden Dataset</h3>
                <div style="display:flex;gap:10px;align-items:center;">
                    <select id="goldenViewSelect" onchange="loadGoldenDataset()" style="padding:6px 12px;border:2px solid rgba(255,255,255,0.5);border-radius:6px;font-size:0.82rem;font-weight:600;color:white;cursor:pointer;background:rgba(255,255,255,0.15);">
                        <option value="nonagentic" style="color:#333;">Non-Agentic</option>
                        <option value="agentic" style="color:#333;">Agentic</option>
                        <option value="evaluation" style="color:#333;">Evaluation Dataset</option>
                    </select>
                    <button class="btn btn-success" onclick="downloadGoldenDataset()" style="background:rgba(255,255,255,0.2);border:1px solid rgba(255,255,255,0.5);color:white;padding:6px 16px;border-radius:6px;cursor:pointer;font-weight:600;">&#11015; Download</button>
                </div>
            </div>
            <div class="card-content">
                <div id="goldenDatasetTable"><p style="color:#888;text-align:center;padding:40px;">No golden dataset available yet. Complete the Augmentation Pipeline to generate one.</p></div>
            </div>
        </div>"""
''')

    # ==================== Update tabs/model_quality_assurance.py path ====================
    # Only the create_disaggregated_table and load_recovery_loop_global use file paths
    # They already accept paths as parameters, so no change needed there.

    # ==================== Update tabs/trustworthy_assurance.py path ====================
    w("tabs/trustworthy_assurance.py", r'''"""Trustworthy Assurance tab - trustworthy table with pagination."""
import json
import pandas as pd
from data_paths import DATA_DIR

try:
    from loguru import logger
except ImportError:
    import logging
    logger = logging.getLogger(__name__)


def generate_trustworthy_table(excel_file, sheet_name, table_id, auto_render=False):
    try:
        excel_path = DATA_DIR / excel_file
        df = pd.read_excel(excel_path, sheet_name=sheet_name)
        if df.empty:
            return "<div class='text-center p-3'>No data available</div>"
        columns = list(df.columns)
        rows = []
        for _, row in df.iterrows():
            cells = []
            for col in columns:
                val = row[col]
                if pd.isna(val) or val is None:
                    cells.append("N/A")
                else:
                    sv = str(val).strip()
                    if str(col).strip().lower() == "result":
                        if sv.lower() == "pass":
                            cells.append("<span style='display:inline-flex;align-items:center;gap:5px;background:#28a745;color:white;padding:4px 12px;border-radius:20px;font-size:0.82rem;font-weight:700;'>&#10003; Pass</span>")
                        elif sv.lower() == "fail":
                            cells.append("<span style='display:inline-flex;align-items:center;gap:5px;background:#dc3545;color:white;padding:4px 12px;border-radius:20px;font-size:0.82rem;font-weight:700;'>&#10007; Fail</span>")
                        else:
                            cells.append(sv)
                    else:
                        cells.append(sv.replace('\n', '<br>').replace('\r', '').replace("'", "&#39;"))
            rows.append(cells)
        rj = json.dumps(rows)
        cl = json.dumps([str(c).strip().lower() for c in columns])
        tid = f"{table_id}Body"; rid = f"{table_id}Rpp"; pid = f"{table_id}PgBtns"; iid = f"{table_id}PgInfo"
        rfn = f"render{table_id[0].upper()}{table_id[1:]}"
        ajs = f'document.addEventListener("DOMContentLoaded", function() {{ render(); }});' if auto_render else ''
        html = ['<div class="table-container"><table class="details-table"><thead><tr>']
        for col in columns:
            al = ' style="text-align:center;"' if str(col).strip().lower() == "result" else ''
            html.append(f'<th{al}>{str(col).replace("_", " ").title()}</th>')
        html.append(f'</tr></thead><tbody id="{tid}"></tbody></table></div>')
        html.append(f'<div class="pagination-container"><div class="pagination-controls"><select id="{rid}" class="rows-selector"><option value="10">10 rows/page</option><option value="25">25</option><option value="50">50</option><option value="100">100</option></select></div><div class="pagination-controls" id="{pid}"></div><div class="page-info"><span id="{iid}">Page 1 of 1</span></div></div>')
        html.append(f"""<script>(function(){{var rows={rj};var colsLower={cl};var curPage=1;function getRpp(){{var el=document.getElementById('{rid}');return el?parseInt(el.value):10;}}function render(){{var rpp=getRpp(),total=rows.length;var tp=Math.max(1,Math.ceil(total/rpp));if(curPage>tp)curPage=tp;var s=(curPage-1)*rpp,e=Math.min(s+rpp,total);var tbody=document.getElementById('{tid}');if(!tbody)return;var h='';for(var i=s;i<e;i++){{h+='<tr>';for(var ci=0;ci<rows[i].length;ci++){{var ir=colsLower[ci]==='result';var ts=ir?' style="text-align:center;"':'';h+='<td'+ts+'>'+rows[i][ci]+'</td>';}}h+='</tr>';}}tbody.innerHTML=h;var info=document.getElementById('{iid}');if(info)info.textContent='Page '+curPage+' of '+tp;var pb=document.getElementById('{pid}');if(pb){{pb.innerHTML='<button class="page-btn" id="{table_id}Prev">Prev</button><span class="page-btn disabled">'+curPage+' / '+tp+'</span><button class="page-btn" id="{table_id}Next">Next</button>';var pv=document.getElementById('{table_id}Prev');var nx=document.getElementById('{table_id}Next');if(curPage<=1)pv.disabled=true;if(curPage>=tp)nx.disabled=true;pv.addEventListener('click',function(){{if(curPage>1){{curPage--;render();}}}});nx.addEventListener('click',function(){{if(curPage<tp){{curPage++;render();}}}});}}}}var rEl=document.getElementById('{rid}');if(rEl)rEl.addEventListener('change',function(){{curPage=1;render();}});window['{rfn}']=render;{ajs}}})();</script>""")
        return '\n'.join(html)
    except Exception as e:
        return f"<div class='text-center p-3'><b>Error loading {sheet_name}: {e}</b></div>"
''')

    # ==================== Update tabs/secondary_llm.py path ====================
    w("tabs/secondary_llm.py", r'''"""Secondary LLM tab - secondary LLM comparison table."""
import json
import pandas as pd

try:
    from loguru import logger
except ImportError:
    import logging
    logger = logging.getLogger(__name__)

from utils import find_column, truncate_text, normalize_query_match
from data_paths import METRICS_TEMPLATE_PATH


def generate_secondary_llm_table(df, metrics):
    try:
        if df.empty:
            return "<div class='text-center p-3'>No details data available</div>"
        html = ['<div class="table-container"><table class="details-table"><thead><tr>']
        html.append('<th style="position:sticky;left:0;z-index:11;min-width:50px;text-align:center;background:linear-gradient(135deg,#2E86AB 0%,#A23B72 100%);">#</th>')
        html.append('<th>Question</th><th>Response</th>')
        for m in metrics:
            html.append(f'<th>Primary_{m}</th><th>Secondary_{m}</th>')
        html.append('</tr></thead><tbody id="secondaryTableBody"></tbody></table></div>')
        html.append(_generate_secondary_pagination_controls())
        html.append(_generate_secondary_llm_javascript(df, metrics))
        return "\n".join(html)
    except Exception as e:
        return f"<div class='alert alert-error'>Error: {e}</div>"


def _generate_secondary_pagination_controls():
    return '<div class="pagination-container"><div class="pagination-controls"><select id="secondaryRowsPerPage" class="rows-selector" onchange="secondaryChangeRowsPerPage()"><option value="10">10 rows/page</option><option value="25">25</option><option value="50">50</option><option value="100">100</option></select></div><div class="pagination-controls" id="secondaryPaginationButtons"></div><div class="page-info"><span id="secondaryPageInfo">Page 1 of 1</span></div></div>'


def _generate_secondary_llm_javascript(df, metrics):
    try:
        td = []
        for idx, row in df.iterrows():
            qs = str(row.get('Query', 'N/A')).replace('\\n', '<br>').replace('\n', '<br>')
            rs = truncate_text(str(row.get('Response', 'N/A')), 100).replace('\\n', '<br>').replace('\n', '<br>')
            qk = normalize_query_match(qs)
            mc = {}
            for m in metrics:
                st = row.get(f"{m}_status", 'Unknown')
                sc = row.get(m, 'N/A')
                fs = f'{sc:.2f}' if pd.notna(sc) and isinstance(sc, (int, float)) else (str(sc) if sc != 'N/A' else 'N/A')
                cc = 'status-warning' if pd.isna(sc) else ('status-passed' if str(st).lower() == 'passed' else ('status-failed' if str(st).lower() == 'failed' else ('status-skipped' if str(st).lower() == 'skipped' else 'status-unknown')))
                mc[m.strip()] = {'value': fs, 'status': st, 'class': cc, 'clickable': True}
            td.append({'index': int(idx), 'query': qs, 'query_key': qk, 'response': rs, 'metrics': mc})

        smm = {}
        try:
            fp = METRICS_TEMPLATE_PATH
            if fp.exists():
                xls = pd.ExcelFile(fp)
                if "Secondary LLM" in xls.sheet_names:
                    sdf = pd.read_excel(xls, sheet_name="Secondary LLM")
                    sdf = sdf.rename(columns=lambda c: str(c).strip())
                    for _, sr in sdf.iterrows():
                        en = str(sr.get("Eval_Name", "")).strip()
                        qv = str(sr.get("Query", "")).strip()
                        if not en: continue
                        mk = en.strip().lower(); qk2 = normalize_query_match(qv)
                        meta = {}
                        for col in sdf.columns:
                            if str(col).strip().lower() in ("eval_name", "trace_id"): continue
                            val = sr.get(col); meta[col] = None if pd.isna(val) else val
                        existing = smm.setdefault(mk, {}).setdefault(qk2, [])
                        if meta.get("Few_Shot_Example"): existing.insert(0, meta)
                        else: existing.append(meta)
        except Exception as e:
            logger.warning(f"Unable to read Secondary LLM sheet: {e}")

        return f"""<script>let secondaryCurrentPage=1,secondaryRowsPerPage=10;const secondaryTableData={json.dumps(td)};const secondaryMetaMap={json.dumps(smm)};let allSecondaryMetrics={json.dumps(metrics)};function renderSecondaryTable(){{const tb=document.getElementById("secondaryTableBody");if(!tb)return;const s=(secondaryCurrentPage-1)*secondaryRowsPerPage;const e=Math.min(s+secondaryRowsPerPage,secondaryTableData.length);let h="";for(let i=s;i<e;i++){{const r=secondaryTableData[i];const rn=(secondaryCurrentPage-1)*secondaryRowsPerPage+(i-s)+1;h+="<tr>";h+=`<td style="position:sticky;left:0;z-index:5;background:#f8f9fa;text-align:center;font-weight:700;color:#555;border-right:2px solid #dee2e6;min-width:50px;">${{rn}}</td>`;h+=`<td>${{r.query}}</td><td>${{r.response}}</td>`;allSecondaryMetrics.forEach(m=>{{const md=r.metrics[m]||{{value:"N/A",class:"status-skipped",clickable:false}};if(md.clickable)h+=`<td class="text-center"><span class="status-cell ${{md.class}}" onclick="openModal(${{r.index}},'${{m}}')">${{md.value}}</span></td>`;else h+=`<td class="text-center"><span class="status-cell status-skipped">N/A</span></td>`;const mm=secondaryMetaMap[m.toLowerCase()]||{{}};const ml=mm[r.query_key]||[];const mt=ml.length>0?ml[0]:null;const et=(mt&&mt.Error_Type!=null&&mt.Error_Type!=="")?String(mt.Error_Type):"none";const bc=et.toLowerCase()!=="none"?"status-failed":"status-passed";const bl=et.toLowerCase()!=="none"?et:"No Violation";h+=`<td class="text-center"><button class="status-cell ${{bc}}" onclick="openSecondaryModal(${{i}},'${{escapeHtml(m)}}')" style="border:none;padding:8px 10px;border-radius:8px;">${{bl}}</button></td>`;}});h+="</tr>";}}tb.innerHTML=h;const tp=Math.ceil(secondaryTableData.length/secondaryRowsPerPage);document.getElementById("secondaryPageInfo").innerText=`Page ${{secondaryCurrentPage}} of ${{tp}}`;const c=document.getElementById("secondaryPaginationButtons");if(c)c.innerHTML=`<button class="page-btn" onclick="secondaryPreviousPage()" ${{secondaryCurrentPage===1?"disabled":""}}>Prev</button><span class="page-btn disabled">${{secondaryCurrentPage}} / ${{tp}}</span><button class="page-btn" onclick="secondaryNextPage()" ${{secondaryCurrentPage===tp?"disabled":""}}>Next</button>`;}}function secondaryNextPage(){{if(secondaryCurrentPage<Math.ceil(secondaryTableData.length/secondaryRowsPerPage)){{secondaryCurrentPage++;renderSecondaryTable();}}}}function secondaryPreviousPage(){{if(secondaryCurrentPage>1){{secondaryCurrentPage--;renderSecondaryTable();}}}}function secondaryChangeRowsPerPage(){{const s=document.getElementById("secondaryRowsPerPage");secondaryRowsPerPage=parseInt(s.value);secondaryCurrentPage=1;renderSecondaryTable();}}function openSecondaryModal(ri,mn){{const modal=document.getElementById('detailModal');if(!modal)return;modal.querySelectorAll('.modal-left .modal-section').forEach(s=>s.style.display='none');const col=modal.querySelector('.collapsible-section');if(col)col.style.display='none';document.getElementById("scoreReasonBlock").style.display="none";document.getElementById("metricContainer").style.display="none";document.getElementById("secondaryContainer").style.display="block";document.getElementById("rcaContainer").style.display="none";document.getElementById("tracebackFields").innerHTML="";document.getElementById("metricSheetFields").innerHTML="";document.getElementById("rcaContainer").innerHTML="";document.getElementById('modalTitle').textContent=mn+' - Secondary LLM Details';document.getElementById('modalSubtitle').textContent='';const ct=document.getElementById('secondaryContainer');ct.innerHTML='';const r=secondaryTableData[ri];const qk=r.query_key;const mm=secondaryMetaMap[mn.toLowerCase()]||{{}};const ml=mm[qk]||[];if(ml.length===0){{const sec=document.createElement('div');sec.className='modal-section';sec.innerHTML='<h4>No metadata</h4><p>N/A</p>';ct.appendChild(sec);}}else{{ml.forEach((mt,idx)=>{{const keys=Object.keys(mt||{{}}).filter(k=>k&&k.toLowerCase()!=='eval_name'&&k.toLowerCase()!=='trace_id');if(keys.length===0){{const sec=document.createElement('div');sec.className='modal-section';sec.innerHTML=`<h4>Row ${{idx+1}}</h4><p>N/A</p>`;ct.appendChild(sec);}}else{{keys.forEach(k=>{{let v=mt[k];const sec=document.createElement('div');sec.className='modal-section';const hd=document.createElement('h4');hd.textContent=String(k).replace(/_/g,' ');sec.appendChild(hd);const cn=document.createElement('div');if(v===null||v===undefined)cn.innerHTML='<p>N/A</p>';else if(typeof v==='object')cn.innerHTML=`<pre>${{escapeHtml(JSON.stringify(v,null,2)).replace(/\\n/g,"<br>")}}</pre>`;else cn.innerHTML=`<p>${{escapeHtml(String(v)).replace(/\\n/g,"<br>")}}</p>`;sec.appendChild(cn);ct.appendChild(sec);}});}}}});}}modal.style.display='block';}}document.addEventListener("DOMContentLoaded",renderSecondaryTable);</script>"""
    except Exception as e:
        return f"<script>console.error('Error: {e}');</script>"
''')

    # ==================== Update run_all.py ====================
    w("run_all.py", r'''#!/usr/bin/env python3
"""GenAI Monitor - Quick Start. Verifies setup and starts the app."""
import os, sys

def check_setup():
    errors, warnings = [], []
    if not os.path.exists("data"):
        errors.append("data/ folder not found!\n  Fix: mkdir data && copy your Excel files into it")
    else:
        if not os.path.exists("data/Metrics_template02.xlsx"):
            errors.append("data/Metrics_template02.xlsx not found (required)")
        for f in ["data/final_eval_results01.xlsx","data/agent_query_augmentations.xlsx","data/Augmentations 2.xlsx"]:
            if not os.path.exists(f): warnings.append(f"{f} not found (optional)")
    if not os.path.exists(".env"): warnings.append(".env not found (needed for Azure OpenAI)")
    missing = []
    for dep in ['flask','pandas','plotly','openpyxl','requests']:
        try: __import__(dep)
        except ImportError: missing.append(dep)
    if missing: errors.append(f"Missing packages: {', '.join(missing)}\n  Fix: pip install {' '.join(missing)}")
    return errors, warnings

def main():
    print("="*60); print("GenAI Monitor - Setup Check"); print("="*60)
    errors, warnings = check_setup()
    if warnings:
        print(f"\nWARNINGS ({len(warnings)}):")
        for w in warnings: print(f"  - {w}")
    if errors:
        print(f"\nERRORS ({len(errors)}):")
        for e in errors: print(f"  - {e}")
        print("\nFix errors and try again."); sys.exit(1)
    print("\nAll checks passed! Starting Flask server...")
    print("="*60)
    from app import app
    app.run(debug=False, host="0.0.0.0", port=5000)

if __name__ == "__main__": main()
''')

    print(f"\n{'='*60}")
    print("DONE! All paths updated to use data/ folder.")
    print(f"{'='*60}")
    print()
    print("Now do this:")
    print()
    print("  1. Create the data folder and copy files:")
    print(f"     mkdir {BASE}\\data")
    print(f"     copy Metrics_template02.xlsx {BASE}\\data\\")
    print(f"     copy final_eval_results01.xlsx {BASE}\\data\\")
    print(f"     copy agent_query_augmentations.xlsx {BASE}\\data\\")
    print(f'     copy "Augmentations 2.xlsx" {BASE}\\data\\')
    print()
    print("  2. Make sure .env is in genai_monitor/:")
    print(f"     copy .env {BASE}\\")
    print()
    print("  3. Run:")
    print(f"     cd {BASE}")
    print("     python run_all.py")
    print()
    print("Final folder structure:")
    print("  genai_monitor/")
    print("  +-- .env")
    print("  +-- app.py")
    print("  +-- run_all.py")
    print("  +-- data/")
    print("  |   +-- Metrics_template02.xlsx")
    print("  |   +-- final_eval_results01.xlsx")
    print("  |   +-- agent_query_augmentations.xlsx")
    print("  |   +-- Augmentations 2.xlsx")
    print("  +-- routes/")
    print("  +-- tabs/")
    print("  +-- templates/")

if __name__ == "__main__":
    build()
